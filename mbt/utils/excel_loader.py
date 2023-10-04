import openpyxl
import os
import constants


# Initialize the Excel file with column headers if it doesn't exist
def initialize_excel_file():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "Date"
    sheet.cell(row=1, column=2).value = "Time"
    sheet.cell(row=1, column=3).value = "PowerAdapter"
    sheet.cell(row=1, column=4).value = "ChargeStatus"
    sheet.cell(row=1, column=5).value = "BatteryCondition"
    sheet.cell(row=1, column=6).value = "BatteryPercent"
    sheet.cell(row=1, column=7).value = "BatteryHealth"
    sheet.cell(row=1, column=8).value = "CycleCount"
    wb.save(constants.EXCEL_FILE_PATH)


def write_to_excel(mac_battery_info: dict) -> None:

    # Open the Excel file
    wb = openpyxl.load_workbook(constants.EXCEL_FILE_PATH)

    # Select the active sheet
    sheet = wb.active

    # Find the next available row in the sheet
    next_row = sheet.max_row + 1

    # Write the data to the Excel sheet
    for index, (key, value) in enumerate(mac_battery_info.items()):
        sheet.cell(row=next_row, column=(index + 1)).value = value

    # Save the changes to the Excel file
    wb.save(constants.EXCEL_FILE_PATH)
    print("Battery information successfully collected!")


def write_battery_info(mac_battery_info):
    if not os.path.isfile(constants.EXCEL_FILE_PATH):
        initialize_excel_file()
    write_to_excel(mac_battery_info)


if __name__ == "__main__":
    pass
